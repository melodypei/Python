# -*- coding: UTF-8 -*-


from  openpyxl import load_workbook

wb = load_workbook(filename='aaa.xlsx')

sheets = wb.sheetnames

print(sheets)#['齐鲁证券', '国金证券', 'Sheet4', 'Sheet5']

print(sheets[1])

ws = wb[sheets[1]]

rows = ws.rows
columns = ws.columns


for row in rows:
    line = [col.value for col in row]
    print(line)


print(ws['C4'].value)
print(ws.cell(row = 2,column = 1).value)

colC = ws['C']
col_range = ws['B:C']
row2 = ws[2]
row_range = ws[1:3]
##print(colC)
##print(col_range)
##print(row2)
##print(row_range)
for col in col_range:
    for row in col:
        print(row.value)
